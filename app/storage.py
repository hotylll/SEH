from __future__ import annotations

import sqlite3
from contextlib import contextmanager
from pathlib import Path
from collections.abc import Iterator
from datetime import datetime, timedelta, timezone
from typing import Any
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

try:  # optional dependency; a stdlib fallback is used when unavailable.
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:  # pragma: no cover - exercised in environments without openpyxl
    Workbook = None
    Font = None
    PatternFill = None

try:  # optional dependency; a stdlib fallback is used when unavailable.
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except ImportError:  # pragma: no cover - exercised in environments without reportlab
    colors = None
    A4 = None
    ParagraphStyle = None
    getSampleStyleSheet = None
    mm = None
    pdfmetrics = None
    UnicodeCIDFont = None
    Paragraph = None
    SimpleDocTemplate = None
    Spacer = None
    Table = None
    TableStyle = None

from app.analysis import calculate_trends, clean_raw_item, content_hash, validate_clean_result
from app.crawler import collect_from_searxng, is_searxng_endpoint
from app.schemas import DataSource, RawItem, ValidationError, utc_now


SCHEMA = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    role TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'enabled',
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS data_sources (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    source_type TEXT NOT NULL,
    endpoint TEXT NOT NULL,
    keywords TEXT,
    schedule TEXT,
    status TEXT NOT NULL DEFAULT 'enabled',
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS collect_tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_id INTEGER NOT NULL,
    task_status TEXT NOT NULL,
    success_count INTEGER NOT NULL DEFAULT 0,
    failed_count INTEGER NOT NULL DEFAULT 0,
    error_message TEXT,
    started_at TEXT NOT NULL,
    finished_at TEXT,
    FOREIGN KEY(source_id) REFERENCES data_sources(id)
);

CREATE TABLE IF NOT EXISTS raw_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_id INTEGER NOT NULL,
    task_id INTEGER NOT NULL,
    title TEXT NOT NULL,
    content TEXT,
    url TEXT,
    author TEXT,
    published_at TEXT,
    raw_hash TEXT UNIQUE NOT NULL,
    created_at TEXT NOT NULL,
    FOREIGN KEY(source_id) REFERENCES data_sources(id),
    FOREIGN KEY(task_id) REFERENCES collect_tasks(id)
);

CREATE TABLE IF NOT EXISTS clean_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    raw_id INTEGER UNIQUE NOT NULL,
    normalized_title TEXT NOT NULL,
    normalized_content TEXT,
    keywords TEXT,
    quality_score REAL NOT NULL,
    created_at TEXT NOT NULL,
    FOREIGN KEY(raw_id) REFERENCES raw_items(id)
);

CREATE TABLE IF NOT EXISTS topic_trends (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    topic TEXT NOT NULL,
    score REAL NOT NULL,
    direction TEXT NOT NULL,
    period_start TEXT NOT NULL,
    period_end TEXT NOT NULL,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS reports (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    report_name TEXT NOT NULL,
    report_type TEXT NOT NULL,
    file_format TEXT NOT NULL DEFAULT 'xlsx',
    file_path TEXT NOT NULL,
    generated_by TEXT NOT NULL,
    generated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS audit_logs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    action TEXT NOT NULL,
    target_type TEXT NOT NULL,
    target_id INTEGER,
    created_at TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_raw_items_source_time ON raw_items(source_id, published_at);
CREATE INDEX IF NOT EXISTS idx_clean_items_quality ON clean_items(quality_score);
CREATE INDEX IF NOT EXISTS idx_topic_trends_topic_period ON topic_trends(topic, period_start, period_end);
"""


MOCK_POOLS: dict[str, list[tuple[str, str, str, int]]] = {
    "news": [
        (
            "{keyword}领域取得重要突破，行业迎来新发展机遇",
            "最新报道显示，{keyword}在实际应用中展现出巨大潜力，多家企业纷纷加大投入。",
            "/news/breakthrough",
            0,
        ),
        (
            "{keyword}市场规模持续扩大，专家看好未来前景",
            "业内专家指出，{keyword}市场在过去一年中增长了30%，预计未来三年仍保持高速增长。",
            "/news/market",
            1,
        ),
        (
            "{keyword}政策利好频出，多方力量竞相布局",
            "政府部门近日发布{keyword}相关扶持政策，引发行业广泛关注与讨论。",
            "/news/policy",
            2,
        ),
        (
            "{keyword}技术标准正式发布，规范化进程加速",
            "国家标准委员会正式发布{keyword}领域技术标准，为行业规范发展指明方向。",
            "/news/standard",
            3,
        ),
    ],
    "forum": [
        (
            "【讨论】{keyword}到底前景如何？来聊聊真实体验",
            "作为从业者想和大家交流一下{keyword}方向的实际感受，有没有前辈能指点一二？",
            "/forum/thread",
            0,
        ),
        (
            "【求助】新手入门{keyword}，应该从哪里学起？",
            "刚接触{keyword}领域，看了很多资料还是一头雾水，求推荐靠谱的学习路径。",
            "/forum/help",
            1,
        ),
        (
            "【分享】{keyword}项目实战踩坑总结，避坑指南",
            "做了大半年{keyword}相关项目，总结了一些常见问题和解决方案，分享给需要的朋友。",
            "/forum/share",
            2,
        ),
        (
            "【投票】{keyword}细分方向哪个最有前景？",
            "A.技术研发 B.产品应用 C.咨询服务 D.教育培训，大家怎么看？欢迎投票讨论。",
            "/forum/poll",
            3,
        ),
    ],
    "api": [
        (
            "数据推送: {keyword}实时指标监测",
            '{{"metric":"{keyword}","value":85.6,"trend":"rising","timestamp":"2026-06-20T10:00:00Z"}}',
            "/api/v2/metrics",
            0,
        ),
        (
            "数据推送: {keyword}周度报告摘要",
            '{{"report_type":"weekly","keyword":"{keyword}","mention_count":1240,"change_pct":15.3}}',
            "/api/v2/reports",
            1,
        ),
        (
            "数据推送: {keyword}关联知识图谱",
            '{{"nodes":[{{"name":"{keyword}"}},{{"name":"政策"}},{{"name":"市场"}}],"edges":[1,2,1,3]}}',
            "/api/v2/graph",
            2,
        ),
        (
            "数据推送: {keyword}舆情情感分析",
            '{{"sentiment":{{"positive":62,"neutral":28,"negative":10}},"total_samples":2450}}',
            "/api/v2/sentiment",
            3,
        ),
    ],
    "csv": [
        (
            "{keyword}_季度数据汇总表",
            "日期,指标,数值,同比\n2026-04-01,{keyword}搜索量,12500,+12%\n2026-05-01,{keyword}搜索量,13800,+15%",
            "/export/quarterly",
            0,
        ),
        (
            "{keyword}_来源分布统计表",
            "来源,占比,数量\n新闻媒体,45%,562\n论坛社区,30%,375\nAPI接口,15%,188\n其他渠道,10%,125",
            "/export/sources",
            1,
        ),
        (
            "{keyword}_地域分布导出表",
            "省份,热度指数,全国排名\n北京,98.5,1\n上海,95.2,2\n广东,92.8,3\n浙江,88.6,4",
            "/export/geo",
            2,
        ),
        (
            "{keyword}_时间趋势导出表",
            "日期,热度值\n2026-06-20,78\n2026-06-21,82\n2026-06-22,79\n2026-06-23,85",
            "/export/timeline",
            3,
        ),
    ],
}


class Repository:
    def __init__(self, db_path: Path | str = "data/app.db") -> None:
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.report_dir = self.db_path.parent / "reports"

    def connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

    @contextmanager
    def session(self) -> Iterator[sqlite3.Connection]:
        conn = self.connect()
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def initialize(self) -> None:
        with self.session() as conn:
            conn.executescript(SCHEMA)
            self._ensure_report_columns(conn)
            if not conn.execute("SELECT 1 FROM users LIMIT 1").fetchone():
                conn.execute(
                    "INSERT INTO users(username, password_hash, role, status, created_at) VALUES(?,?,?,?,?)",
                    ("admin", "demo-password-hash", "admin", "enabled", utc_now()),
                )

    def _ensure_report_columns(self, conn: sqlite3.Connection) -> None:
        columns = {str(row["name"]) for row in conn.execute("PRAGMA table_info(reports)").fetchall()}
        if "file_format" not in columns:
            conn.execute("ALTER TABLE reports ADD COLUMN file_format TEXT NOT NULL DEFAULT 'xlsx'")

    def create_source(self, source: DataSource) -> dict[str, Any]:
        with self.session() as conn:
            cursor = conn.execute(
                """
                INSERT INTO data_sources(name, source_type, endpoint, keywords, schedule, status, created_at)
                VALUES(?,?,?,?,?,?,?)
                """,
                (source.name, source.source_type, source.endpoint, source.keywords, source.schedule, source.status, utc_now()),
            )
            source_id = int(cursor.lastrowid)
            self._audit(conn, "create_source", "data_sources", source_id)
            row = conn.execute("SELECT * FROM data_sources WHERE id = ?", (source_id,)).fetchone()
            return dict(row)

    def list_sources(self, source_type: str | None = None, status: str | None = None) -> list[dict[str, Any]]:
        sql = "SELECT * FROM data_sources WHERE 1=1"
        params: list[Any] = []
        if source_type:
            sql += " AND source_type = ?"
            params.append(source_type)
        if status:
            sql += " AND status = ?"
            params.append(status)
        sql += " ORDER BY id DESC"
        with self.session() as conn:
            return [dict(row) for row in conn.execute(sql, params)]

    def get_source(self, source_id: int) -> dict[str, Any]:
        with self.session() as conn:
            row = conn.execute("SELECT * FROM data_sources WHERE id = ?", (source_id,)).fetchone()
            if row is None:
                raise ValueError(f"source {source_id} not found")
            return dict(row)

    def _generate_mock_items(self, source: dict[str, Any], count: int = 4) -> list[tuple[str, str, str, str]]:
        """根据数据源的 source_type 和 keywords 生成模拟采集数据。"""
        source_type = source.get("source_type", "news")
        keywords_str = source.get("keywords") or ""
        keywords = [k.strip() for k in keywords_str.split(",") if k.strip()]
        if not keywords:
            keywords = ["信息科技"]

        templates = MOCK_POOLS.get(source_type, MOCK_POOLS["news"])
        items: list[tuple[str, str, str, str]] = []
        now = datetime.now(timezone.utc)

        for i, (title_tpl, content_tpl, url_path, days_ago) in enumerate(templates[:count]):
            kw = keywords[i % len(keywords)]
            title = title_tpl.format(keyword=kw)
            content = content_tpl.format(keyword=kw)
            url = f"https://example.com{url_path}-{i + 1}"
            published_at = (now - timedelta(days=days_ago)).isoformat()
            items.append((title, content, url, published_at))

        return items

    def start_collect_task(self, source_id: int) -> dict[str, Any]:
        source = self.get_source(source_id)
        if source.get("status") == "disabled":
            raise ValidationError(f"source {source_id} is disabled")
        # 优先使用 SearXNG 真实爬取，失败时回退到模拟数据
        if is_searxng_endpoint(source.get("endpoint", "")):
            items = collect_from_searxng(source, max_items=8)
        else:
            items = []
        if not items:
            items = self._generate_mock_items(source)
        now = utc_now()
        with self.session() as conn:
            cursor = conn.execute(
                "INSERT INTO collect_tasks(source_id, task_status, started_at) VALUES(?,?,?)",
                (source_id, "running", now),
            )
            task_id = int(cursor.lastrowid)
            success_count = 0
            duplicate_count = 0
            for title, content, url, published_at in items:
                raw = RawItem(
                    source_id=source_id,
                    task_id=task_id,
                    title=title,
                    content=content,
                    url=url,
                    published_at=published_at,
                    author=source["name"],
                )
                if self._insert_raw_and_clean(conn, raw):
                    success_count += 1
                else:
                    duplicate_count += 1
            task_status = "success"
            failed_count = 0
            error_message = None
            if not items:
                task_status = "failed"
                failed_count = 1
                error_message = "未生成任何采集项"
            elif success_count == 0 and duplicate_count < len(items):
                task_status = "failed"
                failed_count = len(items) - duplicate_count
                error_message = "没有有效的采集项"
            elif success_count == 0 and duplicate_count == len(items):
                error_message = "所有项均为重复，已跳过"
            conn.execute(
                """
                UPDATE collect_tasks
                SET task_status = ?, success_count = ?, failed_count = ?, error_message = ?, finished_at = ?
                WHERE id = ?
                """,
                (task_status, success_count, failed_count, error_message, utc_now(), task_id),
            )
            self.rebuild_trends(conn)
            self._audit(conn, "start_collect_task", "collect_tasks", task_id)
        return self.get_task(task_id)

    def get_task(self, task_id: int) -> dict[str, Any]:
        with self.session() as conn:
            row = conn.execute("SELECT * FROM collect_tasks WHERE id = ?", (task_id,)).fetchone()
            if row is None:
                raise ValueError(f"task {task_id} not found")
            return dict(row)

    def list_items(self, keyword: str | None = None, source_id: int | None = None, limit: int = 50) -> list[dict[str, Any]]:
        sql = """
            SELECT raw_items.id, raw_items.source_id, raw_items.title, raw_items.url, raw_items.published_at,
                   clean_items.normalized_title, clean_items.keywords, clean_items.quality_score
            FROM raw_items
            JOIN clean_items ON clean_items.raw_id = raw_items.id
            WHERE 1=1
        """
        params: list[Any] = []
        if keyword:
            sql += " AND (raw_items.title LIKE ? OR raw_items.content LIKE ? OR clean_items.keywords LIKE ?)"
            like = f"%{keyword}%"
            params.extend([like, like, like])
        if source_id:
            sql += " AND raw_items.source_id = ?"
            params.append(source_id)
        sql += " ORDER BY raw_items.published_at DESC LIMIT ?"
        params.append(limit)
        with self.session() as conn:
            return [dict(row) for row in conn.execute(sql, params)]

    def get_item(self, item_id: int) -> dict[str, Any]:
        with self.session() as conn:
            row = conn.execute(
                """
                SELECT raw_items.*, clean_items.normalized_title, clean_items.normalized_content,
                       clean_items.keywords, clean_items.quality_score
                FROM raw_items
                JOIN clean_items ON clean_items.raw_id = raw_items.id
                WHERE raw_items.id = ?
                """,
                (item_id,),
            ).fetchone()
            if row is None:
                raise ValueError(f"item {item_id} not found")
            return dict(row)

    def list_trends(self, limit: int = 20) -> list[dict[str, Any]]:
        with self.session() as conn:
            rows = conn.execute(
                """
                SELECT topic, score, direction, period_start, period_end
                FROM topic_trends
                ORDER BY score DESC, topic ASC
                LIMIT ?
                """,
                (limit,),
            )
            return [dict(row) for row in rows]

    def get_topic_detail(self, topic: str, limit: int = 20) -> dict[str, Any]:
        like = f"%{topic}%"
        with self.session() as conn:
            trends = [
                dict(row)
                for row in conn.execute(
                    """
                    SELECT topic, score, direction, period_start, period_end
                    FROM topic_trends
                    WHERE topic = ?
                    ORDER BY score DESC
                    """,
                    (topic,),
                )
            ]
            rows = conn.execute(
                """
                SELECT raw_items.id, raw_items.source_id, raw_items.title, raw_items.url, raw_items.published_at,
                       clean_items.normalized_title, clean_items.keywords, clean_items.quality_score
                FROM raw_items
                JOIN clean_items ON clean_items.raw_id = raw_items.id
                WHERE raw_items.title LIKE ? OR raw_items.content LIKE ? OR clean_items.keywords LIKE ?
                ORDER BY raw_items.published_at DESC
                LIMIT ?
                """,
                (like, like, like, limit),
            )
            return {"topic": topic, "series": trends, "items": [dict(row) for row in rows]}

    def create_report(
        self,
        report_type: str = "summary",
        file_format: str = "xlsx",
        generated_by: str = "罗元恒",
    ) -> dict[str, Any]:
        report_type = report_type.lower().strip()
        file_format = file_format.lower().strip()
        if report_type not in VALID_REPORT_TYPES:
            raise ValidationError(f"report_type 必须是 {'/'.join(VALID_REPORT_TYPES)} 之一")
        if file_format not in VALID_REPORT_FORMATS:
            raise ValidationError(f"format 必须是 {'/'.join(VALID_REPORT_FORMATS)} 之一")

        self.report_dir.mkdir(parents=True, exist_ok=True)
        generated_at = utc_now()
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        name = f"{PROJECT_REPORT_PREFIX}-{timestamp}-{report_type}.{file_format}"
        path = self.report_dir / name
        with self.session() as conn:
            payload = self._build_report_payload(conn, report_type, generated_by, generated_at)
            self._write_report_file(path, file_format, payload)
            cursor = conn.execute(
                """
                INSERT INTO reports(report_name, report_type, file_format, file_path, generated_by, generated_at)
                VALUES(?,?,?,?,?,?)
                """,
                (name, report_type, file_format, str(path), generated_by, generated_at),
            )
            report_id = int(cursor.lastrowid)
            self._audit(conn, "create_report", "reports", report_id)
            row = conn.execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
            return self._decorate_report(dict(row))

    def list_reports(self) -> list[dict[str, Any]]:
        with self.session() as conn:
            rows = conn.execute(
                """
                SELECT id, report_name, report_type, file_format, file_path, generated_by, generated_at
                FROM reports
                ORDER BY id DESC
                """
            )
            return [self._decorate_report(dict(row)) for row in rows]

    def get_report(self, report_id: int) -> dict[str, Any]:
        with self.session() as conn:
            row = conn.execute(
                """
                SELECT id, report_name, report_type, file_format, file_path, generated_by, generated_at
                FROM reports
                WHERE id = ?
                """,
                (report_id,),
            ).fetchone()
            if row is None:
                raise ValueError(f"report {report_id} not found")
            return self._decorate_report(dict(row))

    def get_report_file(self, report_id: int) -> tuple[Path, dict[str, Any]]:
        report = self.get_report(report_id)
        path = self._resolve_report_path(str(report["file_path"]))
        if not path.exists() or not path.is_file():
            raise ValueError(f"report file {report_id} not found")
        return path, report

    def _decorate_report(self, row: dict[str, Any]) -> dict[str, Any]:
        row["download_url"] = f"/api/v1/reports/{row['id']}/download"
        return row

    def _resolve_report_path(self, file_path: str) -> Path:
        path = Path(file_path)
        if not path.is_absolute():
            path = Path.cwd() / path
        resolved = path.resolve()
        report_root = self.report_dir.resolve()
        if resolved != report_root and report_root not in resolved.parents:
            raise ValueError("report path is outside reports directory")
        return resolved

    def _build_report_payload(
        self,
        conn: sqlite3.Connection,
        report_type: str,
        generated_by: str,
        generated_at: str,
    ) -> dict[str, Any]:
        source_count = int(conn.execute("SELECT COUNT(*) FROM data_sources").fetchone()[0])
        item_count = int(conn.execute("SELECT COUNT(*) FROM raw_items").fetchone()[0])
        task_count = int(conn.execute("SELECT COUNT(*) FROM collect_tasks").fetchone()[0])
        trends = [
            dict(row)
            for row in conn.execute(
                """
                SELECT topic, score, direction, period_start, period_end
                FROM topic_trends
                ORDER BY score DESC, topic ASC
                LIMIT 20
                """
            )
        ]
        item_limit = 200 if report_type == "detail" else 30
        items = [
            dict(row)
            for row in conn.execute(
                """
                SELECT raw_items.id, raw_items.title, raw_items.url, raw_items.published_at,
                       clean_items.keywords, clean_items.quality_score
                FROM raw_items
                JOIN clean_items ON clean_items.raw_id = raw_items.id
                ORDER BY raw_items.published_at DESC
                LIMIT ?
                """,
                (item_limit,),
            )
        ]
        return {
            "report_type": report_type,
            "generated_by": generated_by,
            "generated_at": generated_at,
            "summary": {
                "source_count": source_count,
                "item_count": item_count,
                "task_count": task_count,
                "trend_count": len(trends),
            },
            "trends": trends,
            "items": items,
        }

    def _write_report_file(self, path: Path, file_format: str, payload: dict[str, Any]) -> None:
        if file_format == "xlsx":
            self._write_xlsx_report(path, payload)
            return
        if file_format == "pdf":
            self._write_pdf_report(path, payload)
            return
        raise ValidationError(f"format 必须是 {'/'.join(VALID_REPORT_FORMATS)} 之一")

    def _write_xlsx_report(self, path: Path, payload: dict[str, Any]) -> None:
        if Workbook is None or Font is None or PatternFill is None:
            self._write_basic_xlsx_report(path, payload)
            return

        workbook = Workbook()
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)

        summary_sheet = workbook.active
        summary_sheet.title = "概览"
        summary_sheet.append(["字段", "内容"])
        for cell in summary_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        summary_sheet.append(["报表类型", payload["report_type"]])
        summary_sheet.append(["生成人", payload["generated_by"]])
        summary_sheet.append(["生成时间", payload["generated_at"]])
        for key, value in payload["summary"].items():
            summary_sheet.append([key, value])

        trends_sheet = workbook.create_sheet("趋势榜")
        trends_sheet.append(["主题", "分数", "方向", "周期开始", "周期结束"])
        for cell in trends_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        for trend in payload["trends"]:
            trends_sheet.append(
                [
                    trend["topic"],
                    trend["score"],
                    trend["direction"],
                    trend["period_start"],
                    trend["period_end"],
                ]
            )

        items_sheet = workbook.create_sheet("信息明细")
        items_sheet.append(["ID", "标题", "URL", "发布时间", "关键词", "质量分"])
        for cell in items_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        for item in payload["items"]:
            items_sheet.append(
                [
                    item["id"],
                    item["title"],
                    item["url"],
                    item["published_at"],
                    item["keywords"],
                    item["quality_score"],
                ]
            )

        for sheet in workbook.worksheets:
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)

        workbook.save(path)

    def _write_pdf_report(self, path: Path, payload: dict[str, Any]) -> None:
        reportlab_ready = all(
            (
                colors,
                A4,
                ParagraphStyle,
                getSampleStyleSheet,
                mm,
                pdfmetrics,
                UnicodeCIDFont,
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        )
        if not reportlab_ready:
            self._write_basic_pdf_report(path, payload)
            return

        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        styles = getSampleStyleSheet()
        normal = ParagraphStyle("ChineseNormal", parent=styles["Normal"], fontName="STSong-Light", fontSize=9, leading=13)
        heading = ParagraphStyle("ChineseHeading", parent=styles["Heading1"], fontName="STSong-Light", fontSize=16, leading=22)

        document = SimpleDocTemplate(
            str(path),
            pagesize=A4,
            leftMargin=16 * mm,
            rightMargin=16 * mm,
            topMargin=16 * mm,
            bottomMargin=16 * mm,
        )
        story: list[Any] = [
            Paragraph("集思 · 信息收集整合系统报表", heading),
            Spacer(1, 6),
            Paragraph(f"报表类型：{payload['report_type']}", normal),
            Paragraph(f"生成人：{payload['generated_by']}", normal),
            Paragraph(f"生成时间：{payload['generated_at']}", normal),
            Spacer(1, 8),
        ]

        summary_rows = [["字段", "内容"]] + [[key, str(value)] for key, value in payload["summary"].items()]
        story.append(self._pdf_table(summary_rows, normal, [45 * mm, 90 * mm]))
        story.append(Spacer(1, 10))

        trend_rows = [["主题", "分数", "方向", "周期"]]
        for trend in payload["trends"][:12]:
            trend_rows.append(
                [
                    str(trend["topic"])[:32],
                    str(trend["score"]),
                    str(trend["direction"]),
                    f"{trend['period_start']} ~ {trend['period_end']}",
                ]
            )
        story.append(Paragraph("趋势榜", normal))
        story.append(self._pdf_table(trend_rows, normal, [58 * mm, 22 * mm, 24 * mm, 48 * mm]))
        story.append(Spacer(1, 10))

        item_rows = [["ID", "标题", "发布时间", "质量分"]]
        for item in payload["items"][:18]:
            item_rows.append(
                [
                    str(item["id"]),
                    str(item["title"])[:42],
                    str(item["published_at"])[:19],
                    str(item["quality_score"]),
                ]
            )
        story.append(Paragraph("信息明细", normal))
        story.append(self._pdf_table(item_rows, normal, [14 * mm, 83 * mm, 38 * mm, 18 * mm]))
        document.build(story)

    def _pdf_table(self, rows: list[list[Any]], style: ParagraphStyle, widths: list[float]) -> Table:
        table_rows = [[Paragraph(str(cell), style) for cell in row] for row in rows]
        table = Table(table_rows, colWidths=widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#8FA8BD")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        return table

    def _write_basic_xlsx_report(self, path: Path, payload: dict[str, Any]) -> None:
        sheets = [
            (
                "概览",
                [["字段", "内容"]]
                + [
                    ["报表类型", payload["report_type"]],
                    ["生成人", payload["generated_by"]],
                    ["生成时间", payload["generated_at"]],
                ]
                + [[key, value] for key, value in payload["summary"].items()],
            ),
            (
                "趋势榜",
                [["主题", "分数", "方向", "周期开始", "周期结束"]]
                + [
                    [
                        trend["topic"],
                        trend["score"],
                        trend["direction"],
                        trend["period_start"],
                        trend["period_end"],
                    ]
                    for trend in payload["trends"]
                ],
            ),
            (
                "信息明细",
                [["ID", "标题", "URL", "发布时间", "关键词", "质量分"]]
                + [
                    [
                        item["id"],
                        item["title"],
                        item["url"],
                        item["published_at"],
                        item["keywords"],
                        item["quality_score"],
                    ]
                    for item in payload["items"]
                ],
            ),
        ]
        with ZipFile(path, "w", ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", self._xlsx_content_types(len(sheets)))
            archive.writestr("_rels/.rels", self._xlsx_root_rels())
            archive.writestr("xl/workbook.xml", self._xlsx_workbook(sheets))
            archive.writestr("xl/_rels/workbook.xml.rels", self._xlsx_workbook_rels(len(sheets)))
            for index, (_, rows) in enumerate(sheets, start=1):
                archive.writestr(f"xl/worksheets/sheet{index}.xml", self._xlsx_sheet(rows))

    def _xlsx_content_types(self, sheet_count: int) -> str:
        sheet_overrides = "".join(
            f'<Override PartName="/xl/worksheets/sheet{index}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            for index in range(1, sheet_count + 1)
        )
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            f"{sheet_overrides}</Types>"
        )

    def _xlsx_root_rels(self) -> str:
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
            'Target="xl/workbook.xml"/>'
            "</Relationships>"
        )

    def _xlsx_workbook(self, sheets: list[tuple[str, list[list[Any]]]]) -> str:
        sheet_nodes = "".join(
            f'<sheet name="{escape(name)}" sheetId="{index}" r:id="rId{index}"/>'
            for index, (name, _) in enumerate(sheets, start=1)
        )
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            f"<sheets>{sheet_nodes}</sheets></workbook>"
        )

    def _xlsx_workbook_rels(self, sheet_count: int) -> str:
        rels = "".join(
            f'<Relationship Id="rId{index}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            f'Target="worksheets/sheet{index}.xml"/>'
            for index in range(1, sheet_count + 1)
        )
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f"{rels}</Relationships>"
        )

    def _xlsx_sheet(self, rows: list[list[Any]]) -> str:
        row_nodes = []
        for row_index, row in enumerate(rows, start=1):
            cells = []
            for col_index, value in enumerate(row, start=1):
                cell_ref = f"{self._xlsx_column_name(col_index)}{row_index}"
                text = escape(str(value if value is not None else ""))
                cells.append(f'<c r="{cell_ref}" t="inlineStr"><is><t>{text}</t></is></c>')
            row_nodes.append(f'<row r="{row_index}">{"".join(cells)}</row>')
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f'<sheetData>{"".join(row_nodes)}</sheetData></worksheet>'
        )

    def _xlsx_column_name(self, index: int) -> str:
        name = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            name = chr(65 + remainder) + name
        return name

    def _write_basic_pdf_report(self, path: Path, payload: dict[str, Any]) -> None:
        lines = [
            "Jisi InfoHub Report",
            f"Report type: {payload['report_type']}",
            f"Generated by: {payload['generated_by']}",
            f"Generated at: {payload['generated_at']}",
            f"Sources: {payload['summary']['source_count']}",
            f"Items: {payload['summary']['item_count']}",
            f"Tasks: {payload['summary']['task_count']}",
            f"Trends: {payload['summary']['trend_count']}",
            "Top trends:",
        ]
        for trend in payload["trends"][:8]:
            lines.append(f"- {trend['topic']} / {trend['score']} / {trend['direction']}")
        self._write_minimal_pdf(path, lines)

    def _write_minimal_pdf(self, path: Path, lines: list[str]) -> None:
        text_lines = ["BT", "/F1 11 Tf", "50 800 Td", "14 TL"]
        for line in lines:
            text_lines.append(f"({self._pdf_escape(line)}) Tj")
            text_lines.append("T*")
        text_lines.append("ET")
        stream = "\n".join(text_lines).encode("latin-1", errors="replace")
        objects = [
            b"<< /Type /Catalog /Pages 2 0 R >>",
            b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
            b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
            b"/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
            b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
            b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
        ]
        content = bytearray(b"%PDF-1.4\n")
        offsets = [0]
        for index, obj in enumerate(objects, start=1):
            offsets.append(len(content))
            content.extend(f"{index} 0 obj\n".encode("ascii"))
            content.extend(obj)
            content.extend(b"\nendobj\n")
        xref_offset = len(content)
        content.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
        content.extend(b"0000000000 65535 f \n")
        for offset in offsets[1:]:
            content.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
        content.extend(
            (
                "trailer\n"
                f"<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
                "startxref\n"
                f"{xref_offset}\n"
                "%%EOF\n"
            ).encode("ascii")
        )
        path.write_bytes(bytes(content))

    def _pdf_escape(self, value: Any) -> str:
        text = str(value)
        ascii_text = "".join(ch if 32 <= ord(ch) <= 126 else "?" for ch in text)
        return ascii_text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    def health(self) -> dict[str, Any]:
        with self.session() as conn:
            conn.execute("SELECT 1").fetchone()
            return {
                "status": "ok",
                "db": str(self.db_path),
                "version": "1.0.0",
                "time": utc_now(),
            }

    def seed_demo(self) -> None:
        self.initialize()
        if self.list_sources():
            return
        sources = [
            DataSource(
                name="科技新闻",
                source_type="news",
                endpoint="https://your-searxng-instance.com",
                keywords="人工智能,大模型,AI应用",
                schedule="daily",
            ),
            DataSource(
                name="产业动态",
                source_type="news",
                endpoint="https://your-searxng-instance.com",
                keywords="新能源汽车,半导体,光伏",
                schedule="daily",
            ),
            DataSource(
                name="开源社区",
                source_type="forum",
                endpoint="https://your-searxng-instance.com",
                keywords="Python,开源项目,开发者",
                schedule="weekly",
            ),
            DataSource(
                name="财经资讯",
                source_type="api",
                endpoint="https://your-searxng-instance.com",
                keywords="宏观经济,数字经济,创投",
                schedule="daily",
            ),
        ]
        for source in sources:
            created = self.create_source(source)
            try:
                self.start_collect_task(int(created["id"]))
            except Exception as exc:
                print(f"[seed] 采集失败 ({source.name}): {exc}")

    def rebuild_trends(self, conn: sqlite3.Connection) -> None:
        rows = conn.execute(
            """
            SELECT clean_items.keywords, raw_items.published_at
            FROM clean_items
            JOIN raw_items ON raw_items.id = clean_items.raw_id
            """
        ).fetchall()
        trends = calculate_trends(dict(row) for row in rows)
        # 从数据中自动提取统计周期
        dates = [str(row["published_at"] or utc_now())[:10] for row in rows]
        period_start = min(dates) if dates else utc_now()[:10]
        period_end = max(dates) if dates else utc_now()[:10]
        conn.execute("DELETE FROM topic_trends")
        now = utc_now()
        for trend in trends:
            conn.execute(
                """
                INSERT INTO topic_trends(topic, score, direction, period_start, period_end, created_at)
                VALUES(?,?,?,?,?,?)
                """,
                (trend["topic"], trend["score"], trend["direction"], period_start, period_end, now),
            )

    def _insert_raw_and_clean(self, conn: sqlite3.Connection, raw: RawItem) -> bool:
        raw_hash = content_hash(raw.title, raw.content, raw.url)
        try:
            cursor = conn.execute(
                """
                INSERT INTO raw_items(source_id, task_id, title, content, url, author, published_at, raw_hash, created_at)
                VALUES(?,?,?,?,?,?,?,?,?)
                """,
                (
                    raw.source_id,
                    raw.task_id,
                    raw.title,
                    raw.content,
                    raw.url,
                    raw.author,
                    raw.published_at,
                    raw_hash,
                    utc_now(),
                ),
            )
        except sqlite3.IntegrityError:
            return False
        raw_id = int(cursor.lastrowid)
        clean = validate_clean_result(clean_raw_item(raw))
        conn.execute(
            """
            INSERT INTO clean_items(raw_id, normalized_title, normalized_content, keywords, quality_score, created_at)
            VALUES(?,?,?,?,?,?)
            """,
            (
                raw_id,
                clean["normalized_title"],
                clean["normalized_content"],
                clean["keywords"],
                clean["quality_score"],
                utc_now(),
            ),
        )
        return True

    def _audit(self, conn: sqlite3.Connection, action: str, target_type: str, target_id: int) -> None:
        conn.execute(
            "INSERT INTO audit_logs(action, target_type, target_id, created_at) VALUES(?,?,?,?)",
            (action, target_type, target_id, utc_now()),
        )


PROJECT_REPORT_PREFIX = "jisi-infohub-report"
VALID_REPORT_TYPES = ("summary", "detail")
VALID_REPORT_FORMATS = ("xlsx", "pdf")
